import json
from openpyxl import load_workbook
import os
from PIL import Image
import pyqrcode
# from generate_image import *



workbook = load_workbook(filename="lafia_batch_2_2114.xlsx")
sheet = workbook['LAFIA BATCH 1']

persons = {}

# Using the values_only because you want to return the cells' values
for row in sheet.iter_rows(min_row=2,
                           min_col=2,
                           max_col=22,
                           values_only=True):
    ben_id = row[0]
    person = {
        "app_id": row[1],
        "surname": row[2],
        "middle_name": row[3],
        "first_name": row[4],
        "phone": row[5],
        "address": row[6],
        "lga": row[7],
        "ward": row[8],
        "town": row[9],
        "community": row[10],
        "dob": row[11],
        "status": row[12],
        "sex": row[13],
        "g_name": row[14],
        "g_phone": row[15],
        "e_remark": row[16],
        "f_remark": row[17],
        "pic": row[18],
        "alt_pic": row[19],
        "url": str(row[20]).replace(' ', '')
    }
    if ben_id:
        persons[ben_id] = person

# Using json here to be able to format the output for displaying later
# json_string = json.dumps(persons)

for beneficiary_id in persons:
    if not os.path.isdir(beneficiary_id):
        new_id = beneficiary_id.replace('/', '-')

        qrobj = pyqrcode.create(persons[beneficiary_id]['url'])
        print("Person id: {}, URL: {}".format(beneficiary_id, persons[beneficiary_id]['url']))
        with open(new_id + '.png', 'wb') as f:
            qrobj.png(f, scale=10, module_color='#51c2d5')

        # Now open that png image to put the logo
        img = Image.open(new_id + '.png')
        img = img.convert('RGBA')
        width, height = img.size

        # How big the logo we want to put in the qr code png
        logo_size = 60

        # Open the logo image
        # print(os.getcwd())

        logo = Image.open("logo.png")

        # Calculate xmin, ymin, xmax, ymax to put the logo
        xmin = ymin = int((width / 2) - (logo_size / 2))
        xmax = ymax = int((width / 2) + (logo_size / 2))

        # resize the logo as calculated
        logo = logo.resize((xmax - xmin, ymax - ymin))

        # put the logo in the qr code
        img.paste(logo, (xmin, ymin, xmax, ymax))

        img.save(new_id + '.png')
